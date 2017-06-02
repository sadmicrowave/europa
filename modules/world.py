#!/usr/local/bin/python3


# The parsing method goes through each line of the file and splits the line into cells. 
# Using a double for loop is a common way of working with grids. The x and y variables 
# keep track of the coordinates. The last line is the most interesting, but it`s fine if 
# you don`t fully understand it.

# The variable _world is a dictionary that maps a coordinate pair to a tile. So the code _world[(x, y)] 
# creates the key (i.e. the coordinate pair) of the dictionary. If the cell is an empty string, we dont 
# want to store a tile in it`s place which is why we have the code None if tile_name == ``. However, if 
# the cell does contain a name, we want to actually create a tile of that type. The getattr method is built 
# into Python and lets us reflect into the tile module and find the class whose name matches tile_name. 
# Finally the (x, y) passes the coordinates to the constructor of the tile.

import re, textwrap, copy, sys
from openpyxl import load_workbook

from modules import tiles
from modules import items
from modules import armor
from modules import weapons
from modules import serums
from modules import enemies
from modules import merchants
from modules import repair

class World:
	
	def __init__(self):
		self._world = {}
		self._objects = {}
		self.world_file = 'res/world.xlsx'
		self.start_point = (0,0)

	def load_tiles(self):
		"""Parses a file that describes the world space into the _world object"""
		_rooms = self.construct_room()
		with open('res/map.txt', 'r') as f:
			rows = f.readlines()
		x_max = len(rows[0].split('\t')) # Assumes all rows contain the same number of tabs
    
		for y in range(len(rows)):
			cols = rows[y].split('\t')
			for x in range(x_max):
				tile_name = cols[x].replace('\n', '') # Windows users may need to replace `\r\n`
				if tile_name == 'L1StartingRoom':
					self.start_point = (x,y)
				tile = None if tile_name not in _rooms else _rooms[tile_name]
				if tile:
					tile.x = x
					tile.y = y
					#self._world[(x, y)] = tile
					#self.__dict__[(x,y)] = tile
					self.__dict__['(%s, %s)' % (x,y)] = tile
					
							
	def construct_room(self):
		"""Construct the tile based on the components found in the world spreadsheet for tiles corresponding to the tile name"""
		_rooms = {}
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name('Rooms')
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0 and row[0].value:
				# create a name for the class object name
				room_name = row[0].value.replace(' ','')
				# define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, d, o, i, b):
				#	# initialize super of class type
				#	#print(self.__class__, self)
				#	#super(tiles.MapTile.__class__, tiles.MapTile).__init__(n, d, o, i, b)
				#	super(self.__class__, self).__init__(n, d, o, i, b)
				
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				##room = type(room_name, (eval("tiles.{}".format(row[1].value)),), {'__init__':__constructor__})
				#room = type(row[1].value, (eval("tiles.{}".format(row[1].value)),), {'__init__':__constructor__})
				
				# get the items located in the room by splitting the "Items" column on comma
				room_items = []
				if row[4].value:
					# remove all whitespace in string
					objects = re.sub(r'\s+', '', row[4].value).split(',')
					# iterate over the items in the room
					for item in objects:
						# if the item is an actual world object that has been created
						if item in self._objects:
							# add the item object to the room_items array to pass to the creation of the room tile
							# We append a deepcopy of the original object reference in the _objects dict so we don't
							# override the original object each time we reference it in the future
							room_items.append( copy.deepcopy(self._objects[item]) )
				
				# create item object for interactionable item
				room_interaction_items = []
				if row[5].value:
					# remove all whitespace in string
					objects = re.sub(r'\s+', '', row[5].value).split(',')
					# iterate over the items in the room
					for item in objects:
						# if the item is an actual world object that has been created
						if item in self._objects:
							# add the item object to the room_items array to pass to the creation of the room tile
							# We append a deepcopy of the original object reference in the _objects dict so we don't
							# override the original object each time we reference it in the future
							room_interaction_items.append( copy.deepcopy(self._objects[item]) )
				
				isBlocked	  = row[6].value if row[6].value else ''
				blocked_state = True if str(isBlocked).upper() == 'TRUE' else False
				
				# get the room floor level, must always be found in second character space of every room name
				flevel = room_name[1:2] if room_name[1:2].isdigit() else 1
				
				# add new object to the global _objects object to be used throughout the world
				#_rooms[ room_name ] = room(room_name, textwrap.fill(row[2].value if row[2].value is not None else '',70).strip(), room_items, room_interaction_items, blocked_state)
				_rooms[ room_name ] = eval( "%s.%s(room_name, textwrap.fill(row[2].value if row[2].value is not None else '',70).strip(), textwrap.fill(row[3].value if row[3].value is not None else '',70).strip(), room_items, room_interaction_items, blocked_state, flevel)" % ('tiles',row[1].value) )
		
		return _rooms
							
	def load_base(self,name):
		"""Parse the world spreadsheet tab containing object data for the passed spreadsheet name's items"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				# create a name for the class object name
				item_name = row[0].value.replace(' ','')
				# define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, c, h, l):
				#	# initialize super of class type
				#	#super(self.__class__, self).__init__(name=n, classtype=cl, description=d, cost=c, hp=h, level=l)
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, cost=c, hp=h, level=l)
				#	#print(self.__class__, self)
				#
				#	
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				## add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)' % (name, row[1].value) )
				#self._objects[ item_name ] = items.Item(name=row[0].value, classtype=row[1].value, description=row[2].value, cost=row[3].value, hp=row[4].value, level=row[5].value)
				
	
	def load_armor(self,name='armor'):
		"""Parse the world spreadsheet:Armor tab containing object data for armor items"""
		#self.load_base(name)
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				# create a name for the class object name
				item_name = row[0].value.replace(' ','')
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, c, bl, h, l):
				#	# initialize super of class type
				#	#super(eval('%s.%s' % ('armor',cl)), self).__init__(name=n, classtype=cl, description=d, cost=c, block=bl, hp=h, level=l)
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, cost=c, block=bl, hp=h, level=l)
				#
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				## add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)' % (name,row[1].value) )
				#print( self._objects[ item_name ] )
				

	def load_weapons(self,name='weapons'):
		"""Parse the world spreadsheet:Weapons tab containing object data for weapon items"""
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				# create a name for the class object name
				item_name = row[0].value.replace(' ','')
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, c, da, h, l):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, cost=c, damage=da, hp=h, level=l)
				#
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				# add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)' % (name,row[1].value) )

				#self._objects[ item_name ] = items.Item(name=row[0].value, classtype=row[1].value, description=row[2].value, cost=row[3].value, damage=row[4].value, hp=row[5].value, level=row[6].value)				
					
				
	def load_serums(self,name='serums'):
		"""Parse the world spreadsheet serums tab containing object data for serum items"""
		self.load_base(name)
	
	def load_enemies(self,name='enemies'):
		"""Parse the world spreadsheet Enemies tab containing object data for the passed spreadsheet name's items"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				## create a name for the class object name
				item_name = row[0].value.replace(' ','')
				# define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, h, da, dm, l):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, hp=h, damage=da, dead_message=dm, level=l)
				#
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				## add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value)' % (name, row[1].value) )
	
	
	def load_items(self,name='items'):
		"""Parse the world spreadsheet Items tab containing object data for the passed spreadsheet name's items"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				## create a name for the class object name
				item_name = row[0].value.replace(' ','')
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, c, h, l):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, cost=c, hp=h, level=l)
				#
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				## add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)' % (name, row[1].value) )
				
	
	
	def load_barrier_items(self,name='barriers'):
		"""Parse the world spreadsheet Items tab containing object data for the passed spreadsheet name's barriers"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				## create a name for the class object name
				item_name = row[0].value.replace(' ','')
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, i):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, items=i)
				
				# create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				usable_items = []
				if row[3].value:
					# remove all whitespace in string
					objects = re.sub(r'\s+', '', row[3].value).split(',')
					# iterate over the items in the room
					for item in objects:
						# if the item is an actual world object that has been created
						if item in self._objects:
							# add the item object to the room_items array to pass to the creation of the room tile
							# We append a deepcopy of the original object reference in the _objects dict so we don't
							# override the original object each time we reference it in the future
							usable_items.append( copy.deepcopy(self._objects[item]) )

				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format('items',row[1].value)),), {'__init__':__constructor__})
				## add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value)	
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)' % ('items', row[1].value) )
	
	
	def load_merchants(self,name='merchants'):
		"""Parse the world spreadsheet Merchant tab containing object data for the passed spreadsheet name's items"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				## create a name for the class object name
				item_name = row[0].value.replace(' ','')
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, i):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, items=i)
				#
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#merchant = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				
				# get the items located in the room by splitting the "Items" column on comma
				merchant_items = []
				if row[3].value:
					# remove all whitespace in string
					objects = re.sub(r'\s+', '', row[3].value).split(',')
					# iterate over the items in the room
					for item in objects:
						# if the item is an actual world object that has been created
						if item in self._objects:
							# add the item object to the room_items array to pass to the creation of the room tile
							# We append a deepcopy of the original object reference in the _objects dict so we don't
							# override the original object each time we reference it in the future
							merchant_items.append( copy.deepcopy(self._objects[item]))

				# add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = merchant(row[0].value, row[1].value, row[2].value, merchant_items)
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, merchant_items)' % (name,row[1].value) )

	
	def load_repair(self,name='repair'):
		"""Parse the world spreadsheet Repair tab containing object data for the passed spreadsheet name's items"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				## create a name for the class object name
				item_name = row[0].value.replace(' ','')
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, l):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, level=l)
				#
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#repair = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				
				# add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = repair(row[0].value, row[1].value, row[2].value, row[3].value)
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, row[3].value)' % (name, row[1].value) )

	
	
	
	
	def load_money(self,name='money'):
		"""Parse the world spreadsheet Money tab containing object data for the passed spreadsheet name's items"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				## create a name for the class object name
				item_name = row[0].value.replace(' ','') + str(row[3].value)
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, a, i):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, amt=a, interaction_item=i)
				
				# get the items located in the container by splitting the "Items" column on comma
				#container_object = []
				#if row[3].value:
				#	# remove all whitespace in string
				#	objects = re.sub(r'\s+', '', str(row[3].value)).split(',')
				#	# iterate over the items in the room
				#	for item in objects:
				#		if item.isdigit() and row[1].value == 'Money' :
				#			item = row[0].value + item
				#		# if the item is an actual world object that has been created
				#		print( item )
				#		#elif item in self._objects:
				#		if item in self._objects:
				#			# add the item object to the room_items array to pass to the creation of the room tile
				#			# We append a deepcopy of the original object reference in the _objects dict so we don't
				#			# override the original object each time we reference it in the future
				#			container_object.append( copy.deepcopy(self._objects[item]) )
				
							
				# create item object for interactionable item
				container_interaction_items = []
				if row[4].value:
					# remove all whitespace in string
					objects = re.sub(r'\s+', '', row[4].value).split(',')
					# iterate over the items in the room
					for item in objects:
						# if the item is an actual world object that has been created
						if item in self._objects:
							# add the item object to the room_items array to pass to the creation of the room tile
							# We append a deepcopy of the original object reference in the _objects dict so we don't
							# override the original object each time we reference it in the future
							container_interaction_items.append( copy.deepcopy(self._objects[item]) )
				
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				## add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value, chest_interaction_items)
				self._objects[ item_name ] = eval( 'items.%s(row[0].value, row[1].value, row[2].value, row[3].value, container_interaction_items)' % (row[1].value) )
	
	
	def load_containers(self,name='containers'):
		"""Parse the world spreadsheet Containers tab containing object data for the passed spreadsheet name's items"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				## create a name for the class object name
				item_name = row[0].value.replace(' ','') # + str(row[3].value)
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, a, i):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, amt=a, interaction_item=i)
				
				# get the items located in the container by splitting the "Items" column on comma
				container_object = []
				if row[3].value:
					# remove all whitespace in string
					objects = re.sub(r'\s+', '', str(row[3].value)).split(',')
					# iterate over the items in the room
					for item in objects:
						#if item.isdigit() and row[1].value == 'Money' :
						#	item = row[0].value + item
						# if the item is an actual world object that has been created
						#print( item )
						#elif item in self._objects:
						if item in self._objects:
							# add the item object to the room_items array to pass to the creation of the room tile
							# We append a deepcopy of the original object reference in the _objects dict so we don't
							# override the original object each time we reference it in the future
							container_object.append( copy.deepcopy(self._objects[item]) )
				
							
				# create item object for interactionable item
				container_interaction_items = []
				if row[4].value:
					# remove all whitespace in string
					objects = re.sub(r'\s+', '', row[4].value).split(',')
					# iterate over the items in the room
					for item in objects:
						# if the item is an actual world object that has been created
						if item in self._objects:
							# add the item object to the room_items array to pass to the creation of the room tile
							# We append a deepcopy of the original object reference in the _objects dict so we don't
							# override the original object each time we reference it in the future
							container_interaction_items.append( copy.deepcopy(self._objects[item]) )
				
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format(name,row[1].value)),), {'__init__':__constructor__})
				## add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value, chest_interaction_items)
				self._objects[ item_name ] = eval( 'items.%s(row[0].value, row[1].value, row[2].value, container_object, container_interaction_items)' % (row[1].value) )
	
	
	
	def load_notes(self,name='notes'):
		"""Parse the world spreadsheet Notes tab containing object data for the passed spreadsheet name's items"""
		# open the workbook containing our world contents
		wb = load_workbook( filename = self.world_file )
		# select the specific worksheet to load data from
		sheet = wb.get_sheet_by_name(name.capitalize())
		# iterate over rows in the worksheet
		for index,row in enumerate( sheet.iter_rows() ):
			# skip the header row in the excel sheet
			if index > 0: 
				## create a name for the class object name
				item_name = row[0].value.replace(' ','') #+ str(row[3].value)
				## define an __init__ constructor method to be used when dynamically creating the object class
				#def __constructor__(self, n, cl, d, nar, s, v):
				#	# initialize super of class type
				#	super(self.__class__, self).__init__(name=n, classtype=cl, description=d, narrative=nar, skill=s, value=v)
							
				## create the object class dynamically, utilizing __constructor__ for __init__ method
				#item = type(item_name, (eval("{}.{}".format('items',row[1].value)),), {'__init__':__constructor__})
				## add new object to the global _objects object to be used throughout the world
				#self._objects[ item_name ] = item(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
				self._objects[ item_name ] = eval( '%s.%s(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)' % ('items',row[1].value) )
				
	
	
	def load_wallet(self):
		"""Load the player's purse object.  Done this way because it does not conform to typical object characteristics found in the world spreadsheet"""
		self._objects['Wallet'] = items.Wallet()
	

	def tile_exists(self, x, y):
		#return self.__dict__.get((x, y))
		return self.__dict__.get('(%s, %s)' % (x, y) )


if __name__ == '__main__':

	world = World()
	world.load_armor()
	world.load_weapons()
	world.load_serums()
	world.load_items()
	world.load_notes()
	world.load_barrier_items()
	world.load_enemies()
	world.load_merchants()
	world.load_repair()
	world.load_money()
	world.load_containers()
	world.load_wallet()
	world.load_tiles()
	
	print( dir(world._world[(2,0)]), world._world[(2,0)].__dict__ )
	#print( world._world[(1,0)].__dict__['objects'][0].name, world._world[(1,0)].__dict__['objects'][0].name == 'Torch' )
	
	
	
	